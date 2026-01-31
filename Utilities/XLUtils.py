import openpyxl

class Excel_methods:

    @staticmethod
    def read_excel_data(file_path,sheet_name,row_number,column_number):
        Excel_file=openpyxl.load_workbook(file_path)
        sheet=Excel_file[sheet_name]
        return sheet.cell(row_number,column_number).value

    @staticmethod
    def write_excel_data(file_path,sheet_name,row_number,column_number,data):
        Excel_file=openpyxl.load_workbook(file_path)
        sheet=Excel_file[sheet_name]
        sheet.cell(row_number,column_number).value=data
        Excel_file.save(file_path)
        Excel_file.close()

    @staticmethod
    def get_count_rows(file_path,sheet_name):
        Excel_file=openpyxl.load_workbook(file_path)
        sheet=Excel_file[sheet_name]
        return sheet.max_row
